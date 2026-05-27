#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1.5"]
# ///
"""create_xlsx — build a Microsoft Excel (.xlsx) workbook from a JSON spec or CSV.

Spec schema (JSON):
{
  "sheets": [
    {
      "name":    "Sales",                 # optional, default Sheet1/2/...
      "headers": ["Region", "Q1", "Q2"],  # optional bold+shaded header row
      "rows":    [["West", 10, 12], ["East", 9, 15]],
      "freeze_header": true,              # optional, freezes the header row
      "column_widths": [18, 10, 10],      # optional, per-column width
      "number_formats": {"1": "#,##0", "2": "#,##0"}  # optional, 0-based col -> Excel fmt
    }
  ],
  "accent": "1F4E35"                       # optional header fill hex
}

Single-sheet shortcut from CSV:
  uv run create_xlsx.py --from-csv /workspace/work/scratch/data.csv --output /workspace/work/exports/data.xlsx
  uv run create_xlsx.py --from-csv data.csv --no-header --sheet-name Raw -o out.xlsx

Examples:
  uv run create_xlsx.py --spec /workspace/work/scratch/book.json --output /workspace/work/exports/book.xlsx
  echo '{"sheets":[{"headers":["A","B"],"rows":[[1,2]]}]}' | uv run create_xlsx.py --stdin -o /workspace/work/exports/x.xlsx

Output: single-line JSON -> {"output_path": "...", "format": "xlsx", "bytes": N, "sheets": 1}
Exit codes: 0 ok | 2 usage/spec error | 3 render error
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path


def _die(msg: str, code: int = 2):
    print(f"create_xlsx: {msg}", file=sys.stderr)
    raise SystemExit(code)


def _load_spec(args: argparse.Namespace) -> dict:
    if args.stdin:
        raw = sys.stdin.read()
    elif args.spec_json is not None:
        raw = args.spec_json
    elif args.spec is not None:
        try:
            raw = Path(args.spec).read_text(encoding="utf-8")
        except OSError as e:
            _die(f"cannot read --spec file: {e}")
    else:
        _die("provide one of --spec FILE, --spec-json STR, --stdin, or --from-csv")
    try:
        spec = json.loads(raw)
    except json.JSONDecodeError as e:
        _die(f"spec is not valid JSON: {e}")
    if not isinstance(spec, dict):
        _die("spec must be a JSON object")
    return spec


def _csv_to_spec(csv_path: Path, has_header: bool, sheet_name: str) -> dict:
    if not csv_path.exists():
        _die(f"csv file not found: {csv_path}")
    with csv_path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    if not rows:
        return {"sheets": [{"name": sheet_name, "rows": []}]}
    if has_header:
        return {"sheets": [{"name": sheet_name, "headers": rows[0], "rows": rows[1:], "freeze_header": True}]}
    return {"sheets": [{"name": sheet_name, "rows": rows}]}


def _coerce(value):
    if isinstance(value, str):
        s = value.strip()
        if s and (s.lstrip("-").isdigit()):
            try:
                return int(s)
            except ValueError:
                return value
        try:
            return float(s)
        except (ValueError, TypeError):
            return value
    return value


def _render(spec: dict, output: Path) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    accent = str(spec.get("accent", "1F4E35")).strip().lstrip("#").upper()
    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        _die("spec needs a non-empty `sheets` array")

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for i, sheet in enumerate(sheets):
        if not isinstance(sheet, dict):
            _die(f"sheets[{i}] must be an object")
        ws = wb.create_sheet(title=str(sheet.get("name") or f"Sheet{i + 1}")[:31])
        headers = sheet.get("headers") or []
        rows = sheet.get("rows") or []

        r = 1
        if headers:
            for c, text in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=c, value=str(text))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            r += 1

        for row in rows:
            if not isinstance(row, list):
                _die(f"sheets[{i}].rows must be a list of lists")
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=_coerce(val))
            r += 1

        if sheet.get("freeze_header") and headers:
            ws.freeze_panes = "A2"

        widths = sheet.get("column_widths") or []
        for c, w in enumerate(widths, start=1):
            try:
                ws.column_dimensions[get_column_letter(c)].width = float(w)
            except (TypeError, ValueError):
                pass

        fmts = sheet.get("number_formats") or {}
        for col_str, fmt in fmts.items():
            try:
                col = int(col_str) + 1
            except (TypeError, ValueError):
                continue
            for row_cells in ws.iter_rows(min_col=col, max_col=col):
                for cell in row_cells:
                    if cell.row == 1 and headers:
                        continue
                    cell.number_format = str(fmt)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return {"sheets": len(sheets)}


def main(argv: "list[str] | None" = None) -> int:
    ap = argparse.ArgumentParser(
        prog="create_xlsx",
        description="Build an .xlsx workbook from a JSON spec or a CSV file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    src = ap.add_mutually_exclusive_group()
    src.add_argument("--spec", metavar="FILE", help="Path to a JSON spec file.")
    src.add_argument("--spec-json", metavar="STR", help="Inline JSON spec string.")
    src.add_argument("--stdin", action="store_true", help="Read the JSON spec from stdin.")
    src.add_argument("--from-csv", metavar="FILE", help="Build a single sheet from a CSV file.")
    ap.add_argument("--no-header", action="store_true", help="With --from-csv: treat row 1 as data, not a header.")
    ap.add_argument("--sheet-name", default="Sheet1", help="With --from-csv: sheet name (default Sheet1).")
    ap.add_argument("-o", "--output", required=True, metavar="PATH", help="Output .xlsx path (parent dirs auto-created).")
    args = ap.parse_args(argv)

    output = Path(args.output)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")

    if args.from_csv:
        spec = _csv_to_spec(Path(args.from_csv), not args.no_header, args.sheet_name)
    else:
        spec = _load_spec(args)

    extra = _render(spec, output)
    print(json.dumps({
        "output_path": str(output.resolve()),
        "format": "xlsx",
        "bytes": output.stat().st_size if output.exists() else 0,
        **extra,
    }))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
